"""
Contém a lógica para gerar relatórios por lotes de arquivos.
"""
import asyncio
import datetime
import os
import re
import traceback
import threading
import fitz
import openpyxl
from typing import Dict, Any, Optional, List
from crewai import Crew, Process, Task, Agent, LLM
from DataBaseManager import DataBaseManager
from FolderManager import cleanup_temp_files
from Tasks import create_document_analysis_tasks, create_reporting_tasks

def blocking_download_wrapper(
        db_manager: DataBaseManager,
        file_id: str,
        original_filename: str,
        temp_dir: str,
        batch_number: int
) -> Optional[str]:
    """
    Wrapper para a função de download que será executada em uma thread separada.
    """
    thread_id = threading.get_ident()  # Obtém ID do thread atual
    print(f"        [Thread:{thread_id} | Lote:{batch_number}] Iniciando db_manager.download_file "
          f"para {original_filename}")
    try:
        # Chama a função de download original
        result_path = db_manager.download_file(file_id, original_filename, temp_dir)
        if result_path:
            print(f"        [Thread:{thread_id} | Lote:{batch_number}] db_manager.download_file para "
                  f"{original_filename} retornou sucesso: {result_path}")
        else:
            print(f"        [Thread:{thread_id} | Lote:{batch_number}] db_manager.download_file para "
                  f"{original_filename} retornou falha (None).")
        return result_path  # Retorna o caminho ou None.
    except Exception as thread_e:
        # Captura e loga qualquer erro dentro da execução do thread
        print(f"        [Thread:{thread_id} | Lote:{batch_number}] ERRO DENTRO da thread de download "
              f"para {original_filename}: {thread_e}")
        traceback.print_exc()
        return None  # Indica falha

async def download_file_with_timeout(
        db_manager: DataBaseManager,
        file_id: str,
        original_filename: str,
        temp_dir: str,
        batch_number: int,
        timeout_seconds: int = 180
) -> Optional[str]:
    """
    Baixa um arquivo com timeout, usando uma thread separada.
    """
    print(f"      [Lote {batch_number}] Preparando para executar download em thread: {original_filename}")
    download_wrapper_args = (db_manager, file_id, original_filename, temp_dir, batch_number)
    download_task_wrapper = asyncio.to_thread(blocking_download_wrapper, *download_wrapper_args)
    print(f"      [Lote {batch_number}] Aguardando download de {original_filename} (max {timeout_seconds}s)...")
    try:
        actual_downloaded_path = await asyncio.wait_for(download_task_wrapper, timeout=timeout_seconds)
        return actual_downloaded_path
    except asyncio.TimeoutError:
        print(f"      [Lote {batch_number}] TIMEOUT! Download de {original_filename} excedeu {timeout_seconds} "
              f"segundos.")
        return None

async def verify_file_existence(actual_downloaded_path: str, batch_number: int) -> bool:
    """
    Verifica se o arquivo baixado existe, usando uma thread separada.
    """
    actual_filename = os.path.basename(actual_downloaded_path)
    print(f"      [Lote {batch_number}] Verificando existência do arquivo {actual_filename} em thread...")
    file_exists = await asyncio.to_thread(os.path.exists, actual_downloaded_path)
    print(f"      [Lote {batch_number}] Verificação de existência para {actual_filename} concluída: {file_exists}")
    return file_exists

def extract_and_truncate_content_simple(
        actual_downloaded_path: str,
        actual_filename: str,
        batch_number: int,
        max_chars: int = 100000 # Defina um limite de caracteres razoável
) -> tuple[str, int]:
    """
    Extrai o conteúdo do arquivo e trunca por caracteres se necessário (SEM API).
    Retorna o conteúdo (possivelmente truncado) e a contagem de PALAVRAS.
    """
    file_content = ""
    extraction_error = None
    word_count = 0
    if not os.path.exists(actual_downloaded_path):
        error_msg = f"ERRO CRÍTICO: Arquivo não encontrado em '{actual_downloaded_path}' no momento da extração."
        print(f"      [Lote {batch_number}] {error_msg}")
        return f"ERRO DURANTE A EXTRAÇÃO: {error_msg}", 0

    try:
        if actual_downloaded_path.lower().endswith(".pdf"):
            try:
                with fitz.open(actual_downloaded_path) as doc:
                    file_content = "".join(page.get_text() for page in doc)
                print(f"      [Lote {batch_number}] Extraído texto do PDF: {actual_filename}")
            except Exception as pdf_error:
                extraction_error = f"Erro ao extrair texto do PDF: {pdf_error}"
                print(f"      [Lote {batch_number}] ERRO ao extrair texto do PDF {actual_filename}: {pdf_error}")
        elif actual_downloaded_path.lower().endswith(".xlsx"):
            try:
                workbook = openpyxl.load_workbook(actual_downloaded_path, read_only=True, data_only=True)
                text_parts = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                text_parts.append(str(cell.value))
                file_content = "\n".join(text_parts) # Junta o texto das células
                print(f"      [Lote {batch_number}] Extraído texto do XLSX: {actual_filename}")
            except Exception as xlsx_error:
                extraction_error = f"Erro ao extrair texto do XLSX: {xlsx_error}"
                print(f"      [Lote {batch_number}] ERRO ao extrair texto do XLSX {actual_filename}: {xlsx_error}")
        else: # Assume outros formatos como texto.
            try:
                # Tenta detectar a codificação, com fallback para utf-8 ignore
                try:
                    with open(actual_downloaded_path, "r", encoding='utf-8') as f:
                        file_content = f.read()
                except UnicodeDecodeError:
                    print(f"      [Lote {batch_number}] Falha ao ler {actual_filename} como UTF-8, "
                          f"tentando latin-1...")
                    try:
                        with open(actual_downloaded_path, "r", encoding='latin-1') as f:
                            file_content = f.read()
                    except Exception as read_error_fallback:
                        print(f"      [Lote {batch_number}] Falha ao ler {actual_filename} com latin-1 "
                              f"também: {read_error_fallback}. Usando ignore.")
                        with open(actual_downloaded_path, "r", encoding="utf-8", errors='ignore') as f:
                            file_content = f.read()
                print(f"      [Lote {batch_number}] Lido conteúdo (como texto) de: {actual_filename}")
            except Exception as read_error:
                extraction_error = f"Erro ao ler arquivo como texto: {read_error}"
                print(f"      [Lote {batch_number}] ERRO ao ler {actual_filename} como texto: {read_error}")

        if extraction_error:
            file_content = f"ERRO DURANTE A EXTRAÇÃO DE CONTEÚDO: {extraction_error}"

        # Contagem de palavras simples
        word_count = len(file_content.split())
        print(f"      [Lote {batch_number}] Contagem de palavras (simples): {word_count} para {actual_filename}")

        # Truncamento simples por caracteres
        if len(file_content) > max_chars:
            print(f"      [Lote {batch_number}] ALERTA: Conteúdo ({len(file_content)} chars) excedeu "
                  f"o limite de {max_chars} caracteres. Truncando...")
            original_len = len(file_content)
            file_content = file_content[:max_chars]
            truncation_warning = "\n\n[... CONTEÚDO TRUNCADO DEVIDO AO LIMITE DE CARACTERES ...]"
            # Garante que o aviso caiba
            if len(file_content) + len(truncation_warning) > max_chars:
                file_content = file_content[:max_chars - len(truncation_warning) - 1]
            file_content += truncation_warning
            print(f"      [Lote {batch_number}] Conteúdo truncado para {len(file_content)} caracteres.")

    except Exception as e:
        print(f"      [Lote {batch_number}] ERRO GERAL durante extração/contagem/truncamento simples "
              f"para {actual_filename}: {e}")
        traceback.print_exc()
        file_content = f"ERRO CRÍTICO AO PROCESSAR CONTEÚDO DO ARQUIVO: {e}"
        word_count = 0 # Zera contagem em caso de erro

    return file_content, word_count

async def process_file_in_batch(
        file: Dict[str, str],
        batch_number: int,
        db_manager: DataBaseManager,
        temp_dir:str,
        llm_instance: LLM
) -> Optional[tuple[Task, str]]:
    """
    Processa um único arquivo: baixa, extrai/trunca (simples) e cria tarefa de análise.
    """
    file_id = file.get('id')
    original_filename = file.get('name')
    if not file_id or not original_filename:
        print(f"    [Lote {batch_number}] Erro: Informações inválidas para o arquivo: {file}. Pulando.")
        return None
    print(f"    [Lote {batch_number}] Iniciando processamento para: {original_filename} (ID: {file_id})")
    actual_downloaded_path: Optional[str] = None
    try:
        actual_downloaded_path = await download_file_with_timeout(
            db_manager, file_id, original_filename, temp_dir, batch_number
        )
        if actual_downloaded_path:
            actual_filename = os.path.basename(actual_downloaded_path)
            print(f"      [Lote {batch_number}] Download reportado como sucesso. Caminho: {actual_downloaded_path}")

            file_exists = await verify_file_existence(actual_downloaded_path, batch_number)

            if file_exists:
                print(f"      [Lote {batch_number}] Arquivo verificado: {actual_filename}. "
                      f"Extraindo/Truncando (simples)...")

                # Usa a função simplificada que não depende do llm_manager
                file_content, word_count = extract_and_truncate_content_simple(
                    actual_downloaded_path, actual_filename, batch_number
                )

                # Cria a tarefa de análise CrewAI usando a função de Tasks.py
                # Passa a llm_instance recebida para a função de criação de tarefa
                analysis_tasks_list = create_document_analysis_tasks(
                    file_path=actual_downloaded_path, # Passa o caminho para a descrição
                    llm=llm_instance # Passa a instância crewai.LLM
                )

                if analysis_tasks_list:
                    analysis_task = analysis_tasks_list[0] # Pega a primeira (e única) tarefa criada
                    # Atualiza a descrição para incluir o conteúdo real (se necessário, ou confia que a
                    # task usará as tools)
                    analysis_task.description = (
                        f"Analise o seguinte conteúdo extraído do documento '{actual_filename}' "
                        f"(originalmente '{original_filename}', Lote {batch_number}). "
                        f"{'O conteúdo PODE TER SIDO TRUNCADO' if not file_content.startswith('ERRO') else 'HOUVE UM ERRO NA EXTRAÇÃO/LEITURA DO CONTEÚDO'}. "
                        f"({word_count} palavras estimadas).\n"
                        "Sua análise DEVE focar em:\n"
                        "1. Se o conteúdo NÃO começar com 'ERRO': Use a ferramenta 'count_words' para contar as palavras no texto fornecido.\n"
                        "2. Se o conteúdo NÃO começar com 'ERRO': Gere um resumo EXTREMAMENTE CONCISO (5 a 10 linhas) do conteúdo principal.\n"
                        "3. Se o conteúdo COMEÇAR com 'ERRO': Retorne a mensagem de erro como 'summary' e 'None' como 'word_count'.\n"
                        "NÃO inclua o texto original na sua resposta final (exceto se for a mensagem de erro). Retorne APENAS o dicionário Python.\n"
                        f"\nCONTEÚDO PARA ANÁLISE:\n---\n{file_content}\n---"
                    )
                    print(f"      [Lote {batch_number}] Tarefa de análise para {actual_filename} criada/atualizada.")
                    return analysis_task, actual_downloaded_path
                else:
                    print(f"      [Lote {batch_number}] ERRO: create_document_analysis_tasks não retornou tarefas.")
                    return None
            else:
                print(f"      [Lote {batch_number}] ERRO INESPERADO: Download reportado como sucesso, "
                      f"mas arquivo não encontrado em: {actual_downloaded_path}")
                return None
        else:
            print(f"      [Lote {batch_number}] Download falhou ou não concluído para: {original_filename}")
            return None

    except Exception as e:
        print(f"      [Lote {batch_number}] Erro EXCEPCIONAL durante processamento (fora download) "
              f"do arquivo {original_filename}: {e}")
        traceback.print_exc()
        # Limpeza do arquivo baixado
        if actual_downloaded_path and os.path.exists(actual_downloaded_path):
            actual_filename = os.path.basename(actual_downloaded_path)
            try:
                os.remove(actual_downloaded_path)
                print(f"      [Lote {batch_number}] Arquivo baixado {actual_filename} removido devido a erro.")
            except OSError as rm_err:
                print(f"      [Lote {batch_number}] Erro ao tentar remover arquivo {actual_filename} após "
                      f"falha: {rm_err}")
        return None


def _handle_gather_exception(result: BaseException, batch_number: int):
    """
    Handles exceptions returned directly by asyncio.gather.
    """
    print(f"   [Lote {batch_number}] Erro capturado pelo asyncio.gather: {result!r}")
    if isinstance(result, Exception):  # Loga traceback para Exceptions "reais".
        traceback.print_exc()

def _validate_and_collect_result(
        result: tuple,
        batch_number: int,
        batch_analysis_tasks: list,
        batch_downloaded_files: list
):
    """
    Validates and collects a single result from asyncio.gather.
    """
    analysis_task, downloaded_file_path = result
    # Verifica se ambos os elementos da tupla são válidos.
    if (analysis_task and isinstance(analysis_task, Task) and downloaded_file_path and
            isinstance(downloaded_file_path, str)):
        batch_analysis_tasks.append(analysis_task)
        batch_downloaded_files.append(downloaded_file_path)
        print(f"     - Resultado válido coletado para: {os.path.basename(downloaded_file_path)}")
    else:
        # Caso onde process_file_in_batch retornou (None, None) ou algo inválido na tupla.
        print(f"   [Lote {batch_number}] Processamento de arquivo não retornou tarefa/caminho válidos "
              f"(recebido: {result!r}).")

def _handle_unexpected_result(result: Any, batch_number: int):
    """
    Handles unexpected result types from asyncio.gather.
    """
    print(f"   [Lote {batch_number}] Resultado inesperado recebido do processamento paralelo: {result!r}")


def process_batch_results(results_from_gather: list, batch_number: int):
    """
    Processa os resultados brutos do asyncio.gather para um lote.
    Separa tarefas bem-sucedidas, arquivos baixados e lida com erros.
    Return:
        tuple[List[Task], List[str]]: Uma tupla contendo a lista de tarefas de análise válidas
                                      e a lista de caminhos de arquivos baixados com sucesso.
    """
    batch_analysis_tasks = []
    batch_downloaded_files = []
    print(f"   [Lote {batch_number}] Processamento paralelo concluído. Coletando e validando resultados...")

    for result in results_from_gather:
        if isinstance(result, BaseException):
            _handle_gather_exception(result, batch_number)
            continue
        if result and isinstance(result, tuple) and len(result) == 2:
            _validate_and_collect_result(result, batch_number, batch_analysis_tasks, batch_downloaded_files)
        elif result is not None:
            _handle_unexpected_result(result, batch_number)

    print(f"   [Lote {batch_number}] Coleta finalizada. {len(batch_analysis_tasks)} tarefas válidas, "
          f"{len(batch_downloaded_files)} arquivos baixados.")
    return batch_analysis_tasks, batch_downloaded_files

def create_report_task(
        batch_number: int,
        analysis_tasks: List[Task],
        llm_instance: LLM,
        specific_report_dir: str
):
    """
    Cria a tarefa final de relatório para um lote, usando a instância LLM fornecida e o diretório de relatório
    específico.
    """
    if not analysis_tasks:
        print(f"   [Lote {batch_number}] Nenhuma tarefa de análise para incluir no relatório.")
        return None

    print(f"   [Lote {batch_number}] Criando tarefa de relatório com {len(analysis_tasks)} análises como contexto...")

    # Usa a função de Tasks.py para criar a tarefa base
    reporting_tasks_list = create_reporting_tasks(
        llm=llm_instance
    )

    if reporting_tasks_list:
        report_task = reporting_tasks_list[0]
        # Define o contexto da tarefa de relatório como as tarefas de análise concluídas
        report_task.context = analysis_tasks

        # Injeta dinamicamente o caminho específico na descrição e expected_output
        original_description = report_task.description
        original_expected_output = report_task.expected_output

        report_task.description = original_description.format(
            specific_report_dir_placeholder=specific_report_dir
        )
        report_task.expected_output = original_expected_output.format(
            specific_report_dir_placeholder=specific_report_dir
        )
        print(f"   [Lote {batch_number}] Tarefa de relatório criada. Diretório de destino: {specific_report_dir}")
        return report_task
    else:
        print(f"   [Lote {batch_number}] ERRO: create_reporting_tasks não retornou tarefas.")
        return None

async def run_batch_crew(
        batch_number: int,
        analysis_tasks: List[Task],
        report_task: Optional[Task],
        document_agent: Any,
        reporting_agent: Any
) -> Optional[str]:
    """
    Executa a CrewAI para um lote específico, incluindo tarefas de análise e a tarefa de relatório.
    O kickoff síncrono é executado em um thread separado.
    Return:
        Optional[str]: A string de resultado da tarefa final (report_task), que geralmente é
                       a confirmação/erro da ferramenta de salvar, ou None se falhar.
    """
    if not analysis_tasks and not report_task:
        print(f"  [Lote {batch_number}] Nenhuma tarefa válida para executar no Crew.")
        return None # Não executa o Crew se não houver tarefas

    tasks_for_crew = analysis_tasks + ([report_task] if report_task else [])
    print(f"  [Lote {batch_number}] Criando Crew com {len(tasks_for_crew)} tarefas ({len(analysis_tasks)} análise(s), "
          f"{'1 relatório' if report_task else 'sem relatório'})...")

    # Definindo os agentes que participarão.
    # Garante que ambos os agentes estejam na lista se houver tarefas para eles
    agents_for_crew = []
    if analysis_tasks:
        agents_for_crew.append(document_agent)
    if report_task:
        # Evita adicionar o mesmo agente duas vezes se ele fizesse as duas tarefas
        if reporting_agent not in agents_for_crew:
            agents_for_crew.append(reporting_agent)

    if not agents_for_crew:
        print(f"  [Lote {batch_number}] ERRO: Nenhuma tarefa ou nenhum agente definido para o Crew.")
        return f"--- Relatório do Lote {batch_number} (ERRO: SEM AGENTES PARA CREW) ---"


    batch_crew = Crew(
        agents=agents_for_crew,
        tasks=tasks_for_crew,
        process=Process.sequential, # Executa tarefas em ordem
        verbose=False, # Manten False para menos ruído, True para debug
        memory=False
    )

    kickoff_result_str = None
    try:
        print(f"  [Lote {batch_number}] Executando Crew.kickoff() em thread separada...")
        # O resultado da crew será o resultado da última tarefa na lista (report_task, se existir)
        kickoff_result = await asyncio.to_thread(batch_crew.kickoff)
        print(f"  [Lote {batch_number}] Execução do Crew.kickoff() concluída.")

        # Converte o resultado para string para tratamento uniforme
        if kickoff_result is not None:
            kickoff_result_str = str(kickoff_result)
        else:
            kickoff_result_str = f"--- Relatório do Lote {batch_number} (RETORNO VAZIO/NULO DA CREW) ---"

        # Log do resultado bruto recebido
        print(f"    [Lote {batch_number}] Resultado bruto do kickoff (tipo {type(kickoff_result)}):"
              f" '{kickoff_result_str}'")

        if report_task:
            print(f"    [Lote {batch_number}] Tarefa de relatório foi executada. Resultado será processado.")
        else:
            print(f"    [Lote {batch_number}] Crew executada apenas para análise.")


    except Exception as e:
        print(f"\n  [Lote {batch_number}] ERRO CRÍTICO durante a execução do Crew.kickoff(): {e}")
        traceback.print_exc()
        # Registra o erro para o relatório final
        kickoff_result_str = f"--- Relatório do Lote {batch_number} (ERRO NA EXECUÇÃO DA CREW: {e}) ---"

    return kickoff_result_str # Retorna a string resultante (confirmação, erro ou None)

def save_final_report(all_reports_content: List[str], report_dir: str):
    """
    Consolida todos os conteúdos de relatórios (strings) em um único arquivo final
    dentro do diretório específico da execução fornecido.
    """
    print(f"\n--- Consolidação e Salvamento do Relatório Final em '{report_dir}' ---")
    if not all_reports_content:
        print("Nenhum conteúdo de relatório parcial foi gerado para consolidar.")
        return
    if not os.path.exists(report_dir):
        print(f"AVISO: O diretório de relatório final '{report_dir}' não existe. Tentando criar...")
        try:
            os.makedirs(report_dir, exist_ok=True)
        except OSError as e:
            print(f"Erro ao criar o diretório final '{report_dir}': {e}. Não é possível salvar o relatório.")
            traceback.print_exc()
            return

    print(f"Combinando {len(all_reports_content)} seções de relatório...")
    # Combina os conteúdos com um separador em Markdown
    final_report_text = "\n\n---\n\n".join(all_reports_content) # Usar '---' para separador horizontal MD
    # Salva o relatório final combinado.
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    report_filename = f"FINAL_consolidated_drive_report.md" # Salva como Markdown.
    report_filepath = os.path.join(report_dir, report_filename) # Usa o report_dir (subpasta específica)
    try:
        with open(report_filepath, "w", encoding="utf-8") as f:
            # Adiciona um título H1 e talvez a data/hora da geração do consolidado
            f.write(f"# Relatório Consolidado de Análise de Documentos\n")
            f.write(f"*Gerado em: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*\n\n")
            f.write(final_report_text)
        print(f"Relatório final consolidado salvo em: {report_filepath}")
    except IOError as e:
        print(f"Erro ao salvar o relatório final consolidado em {report_filepath}: {e}")
        traceback.print_exc()
    except Exception as e:
        print(f"Erro inesperado ao salvar o relatório final: {e}")
        traceback.print_exc()

def extract_report_path(confirmation_string: str) -> Optional[str]:
    """Extrai o caminho do arquivo da string de confirmação da ferramenta."""
    if not confirmation_string or not isinstance(confirmation_string, str):
        return None
    # Tenta encontrar um padrão como "saved to: <caminho>"
    match = re.search(r"saved to:\s*(.*)", confirmation_string, re.IGNORECASE)
    if match:
        path = match.group(1).strip()
        # Verifica se o caminho parece válido (simplista)
        if path.endswith(".txt") and os.path.sep in path:
            # Tenta normalizar o caminho
            try:
                return os.path.normpath(path)
            except Exception:
                return path # Retorna o caminho como encontrado se a normalização falhar
    return None # Retorna None se não encontrar um caminho

# Função principal de orquestração de lotes.
async def process_batches(
        files: List[Dict[str, str]],
        total_files: int,
        total_batches: int,
        batch_size: int,
        db_manager: DataBaseManager,
        document_agent: Agent,
        reporting_agent: Agent,
        temp_dir: str,
        report_dir: str,
        llm_instance: LLM,
):
    """
    Orquestra o processamento dos arquivos em lotes com a instância LLM fornecida,
    salvando relatórios no diretório específico fornecido.
    """
    all_final_reports_content = []
    all_downloaded_files_ever = set()

    for i in range(0, total_files,  batch_size):
        start_index = i
        end_index = min(i + batch_size, total_files)
        current_batch_files = files[start_index:end_index]
        batch_number = (i // batch_size) + 1
        print(f"\n--- Iniciando Lote {batch_number} de {total_batches} ({len(current_batch_files)} arquivos) ---")

        batch_coroutines = []
        print(f"  [Lote {batch_number}] Agendando processamento paralelo de arquivos...")
        for file_info in current_batch_files:
            # Passa llm_instance para process_file_in_batch
            coro = process_file_in_batch(
                file_info, batch_number, db_manager, temp_dir, llm_instance
            )
            batch_coroutines.append(coro)

        # Espera o processamento paralelo (download, extração, criação de task de análise)
        gather_results = await asyncio.gather(*batch_coroutines, return_exceptions=True)

        # Processa os resultados para obter tarefas de análise válidas e arquivos baixados
        batch_analysis_tasks, batch_downloaded_files = process_batch_results(gather_results, batch_number)
        all_downloaded_files_ever.update(batch_downloaded_files) # Adiciona ao set global

        # Cria a tarefa de relatório para o lote, passando llm_instance e o diretório específico
        batch_report_task = create_report_task(
            batch_number, batch_analysis_tasks, llm_instance, report_dir # Passa report_dir (subpasta)
        )

        # Executa a Crew com as tarefas de análise e a tarefa de relatório
        # A crew usará o reporting_agent para a batch_report_task
        report_content_from_crew = await run_batch_crew(
            batch_number,
            batch_analysis_tasks,
            batch_report_task,
            document_agent, # Agente para tarefas de análise
            reporting_agent # Agente para tarefa de relatório
        )

        # Coleta o conteúdo do relatório gerado pela CREW (que é o resultado da última task, a de relatório)
        if batch_report_task: # Só adiciona algo se uma tarefa de relatório foi criada
            if report_content_from_crew:
                # Adiciona a saída da task de relatório (confirmação/caminho ou erro)
                all_final_reports_content.append(
                    f"--- Resultado da Tarefa de Relatório do Lote {batch_number} ---\n"
                    f"{report_content_from_crew}\n"
                    f"----------------------------------------------------------"
                )
            else:
                # Caso a crew tenha rodado mas retornado None/vazio para a task de relatório
                all_final_reports_content.append(
                    f"--- Relatório do Lote {batch_number} ---\n"
                    f"AVISO: A tarefa de relatório foi executada, mas não retornou confirmação ou caminho do "
                    f"arquivo.\n"
                    f"Verifique o diretório: {report_dir}\n"
                    f"----------------------------------------------------------"
                )
        elif batch_analysis_tasks: # Se houve análise mas o relatório não foi nem criado/tentado
            all_final_reports_content.append(f"--- Análise do Lote {batch_number} (Sem Relatório) ---\n"
                                             f"Tarefas de análise concluídas, mas a tarefa de geração de relatório "
                                             f"não foi criada.\n"
                                             f"----------------------------------------------------------")
        # Se nem análise houve, não adiciona nada

        # Limpa os arquivos temporários deste lote (baixados)
        print(f"  [Lote {batch_number}] Iniciando limpeza de arquivos temporários do lote...")
        cleanup_temp_files(batch_downloaded_files, temp_dir) # Limpa de TEMP_DIR
        print(f"--- Lote {batch_number} Finalizado ---")

    # Salva o relatório final consolidado (contendo as confirmações/erros de cada lote)
    save_final_report(all_final_reports_content, report_dir) # Passa report_dir (subpasta)

    print("Processamento de todos os lotes concluído.")